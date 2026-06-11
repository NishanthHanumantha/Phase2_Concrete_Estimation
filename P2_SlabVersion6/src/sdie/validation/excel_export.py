"""Export SDIE pipeline results to Excel."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore
    Font = None  # type: ignore
    PatternFill = None  # type: ignore
    get_column_letter = None  # type: ignore

DEFAULT_REVIEW_THRESHOLD = 75.0
FILL_REVIEW = PatternFill("solid", fgColor="FFF2CC") if PatternFill else None  # amber <75%
FILL_FORCE_REVIEW = PatternFill("solid", fgColor="FCE4D6") if PatternFill else None  # orange <60%


def _require_openpyxl() -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required: pip install openpyxl")


def _bounds_to_dims_m(bounds_cm: list[float] | tuple[float, ...] | None) -> tuple[float | None, float | None]:
    if not bounds_cm or len(bounds_cm) < 4:
        return None, None
    length_m = abs(bounds_cm[2] - bounds_cm[0]) / 100.0
    breadth_m = abs(bounds_cm[3] - bounds_cm[1]) / 100.0
    return round(length_m, 3), round(breadth_m, 3)


def _autosize_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def _highlight_row(ws, row_idx: int, max_col: int, fill) -> None:
    if fill is None:
        return
    for col in range(1, max_col + 1):
        ws.cell(row=row_idx, column=col).fill = fill


def _review_fill(confidence: float | None, *, review_threshold: float, force_threshold: float):
    if confidence is None:
        return None
    if confidence < force_threshold:
        return FILL_FORCE_REVIEW
    if confidence < review_threshold:
        return FILL_REVIEW
    return None


def _load_json_if_exists(path: Path | str | None) -> dict[str, Any] | None:
    if not path:
        return None
    p = Path(path)
    if not p.is_file():
        return None
    return json.loads(p.read_text(encoding="utf-8"))


def _load_export_context(result: dict[str, Any], output_path: Path) -> dict[str, Any]:
    notes = result.get("detection_notes") or {}
    stem = Path(str(result.get("source_dxf", output_path.stem))).stem
    out_dir = output_path.parent

    review_data = _load_json_if_exists(notes.get("review_queue"))
    if review_data is None:
        review_data = _load_json_if_exists(out_dir / f"{stem}_review_queue.json")

    model_data = _load_json_if_exists(out_dir / f"{stem}_building_model.json")
    components: list[dict[str, Any]] = []
    if model_data:
        building = model_data.get("building") or model_data
        components = list(building.get("components") or [])

    thresholds = (review_data or {}).get("thresholds") or {}
    review_threshold = float(thresholds.get("review", DEFAULT_REVIEW_THRESHOLD))
    force_threshold = float(thresholds.get("force_queue", 60.0))

    return {
        "review_queue": review_data,
        "components": components,
        "review_threshold": review_threshold,
        "force_threshold": force_threshold,
    }


def _centroid_in_bounds_cm(centroid_mm: list[float] | None, bounds_cm: list[float]) -> bool:
    if not centroid_mm or len(centroid_mm) < 2 or len(bounds_cm) < 4:
        return False
    x_cm = float(centroid_mm[0]) / 10.0
    y_cm = float(centroid_mm[1]) / 10.0
    return bounds_cm[0] <= x_cm <= bounds_cm[2] and bounds_cm[1] <= y_cm <= bounds_cm[3]


def _slab_classification_stats(
    slab: dict[str, Any],
    components: list[dict[str, Any]],
    *,
    review_threshold: float,
) -> dict[str, Any]:
    bounds = slab.get("bounds_cm") or []
    in_slab = [
        c
        for c in components
        if c.get("component_type") == "Slab"
        and _centroid_in_bounds_cm(c.get("centroid_mm"), bounds)
    ]
    confs = [float(c.get("confidence") or 0) for c in in_slab]
    low_count = sum(1 for v in confs if v < review_threshold)

    thk_conf_pct = round(float(slab.get("thickness_confidence") or 1.0) * 100.0, 1)
    entity_conf = round(min(confs), 1) if confs else None
    thickness_uncertain = slab.get("thickness_source") == "default_note"
    thickness_low_conf = thk_conf_pct < review_threshold
    review_required = (
        low_count > 0
        or thickness_uncertain
        or thickness_low_conf
        or (entity_conf is not None and entity_conf < review_threshold)
    )
    display_conf = entity_conf if entity_conf is not None else thk_conf_pct
    return {
        "entity_confidence_pct": display_conf,
        "low_conf_entity_count": low_count,
        "review_required": review_required,
    }


def _write_summary_sheet(wb, result: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)
    ws["A1"] = "SDIE Pipeline Results"
    ws["A1"].font = bold

    totals = result.get("totals") or {}
    notes = result.get("detection_notes") or {}
    cls = notes.get("classification") or {}
    benchmark = result.get("benchmark") or {}
    inf = result.get("inference_metrics") or {}
    inf_cls = inf.get("classification") or {}
    config = result.get("config") or {}

    rows: list[tuple[str, Any]] = [
        ("Source DXF", Path(str(result.get("source_dxf", ""))).name),
        ("Processed at", result.get("processed_at")),
        ("Engine version", result.get("version")),
        ("Project ID", config.get("project_id")),
        ("Detection mode", config.get("detection_mode")),
        ("", ""),
        ("Slab count", totals.get("slab_count")),
        ("Slab area (m2)", totals.get("area_m2")),
        ("Slab concrete (m3)", totals.get("slab_concrete_m3")),
        ("Beam count", totals.get("beam_count")),
        ("Beam total length (m)", totals.get("beam_total_length_m")),
        ("Beam concrete (m3)", totals.get("beam_concrete_m3")),
        ("Total concrete (m3)", totals.get("concrete_m3")),
        ("Total shuttering (m2)", totals.get("shuttering_m2")),
        ("", ""),
        ("Entities classified", notes.get("entity_count")),
        ("Review required", notes.get("review_required_count")),
        ("Low confidence %", notes.get("low_confidence_pct")),
        ("DeepSeek ambiguous", cls.get("ambiguous_count")),
        ("DeepSeek updated", (cls.get("deepseek") or {}).get("updated") if isinstance(cls.get("deepseek"), dict) else None),
        ("", ""),
        ("Benchmark status", benchmark.get("status")),
        ("Benchmark accuracy %", benchmark.get("overall_accuracy_pct")),
        ("", ""),
        ("Slab+Beam entity share %", inf_cls.get("slab_beam_share_pct")),
        ("Mean confidence Slab+Beam %", inf_cls.get("mean_confidence_slab_beam")),
        ("Low confidence %", inf_cls.get("low_confidence_pct")),
        ("Unknown %", inf_cls.get("unknown_pct")),
        ("Quality checks pass", inf.get("quality_pass")),
    ]

    for idx, (label, value) in enumerate(rows, start=3):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=value)
    ws["A3"].font = bold
    _autosize_columns(ws)


def _write_slabs_sheet(
    wb,
    slabs: list[dict[str, Any]],
    *,
    components: list[dict[str, Any]] | None = None,
    review_threshold: float = DEFAULT_REVIEW_THRESHOLD,
    force_threshold: float = 60.0,
) -> None:
    ws = wb.create_sheet("Slabs")
    headers = [
        "S.No",
        "Slab ID",
        "Strategy",
        "Length (m)",
        "Breadth (m)",
        "Thickness (mm)",
        "Area (m2)",
        "Concrete (m3)",
        "Shuttering (m2)",
        "Thickness Source",
        "Min Slab Entity Conf %",
        "Low Conf Entities",
        "Review",
        "Centroid X (cm)",
        "Centroid Y (cm)",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    components = components or []
    for row_idx, slab in enumerate(slabs, start=2):
        length_m, breadth_m = _bounds_to_dims_m(slab.get("bounds_cm"))
        centroid = slab.get("centroid_cm") or [None, None]
        stats = _slab_classification_stats(
            slab,
            components,
            review_threshold=review_threshold,
        )
        entity_conf = stats["entity_confidence_pct"]
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=slab.get("slab_id"))
        ws.cell(row=row_idx, column=3, value=slab.get("strategy"))
        ws.cell(row=row_idx, column=4, value=length_m)
        ws.cell(row=row_idx, column=5, value=breadth_m)
        ws.cell(row=row_idx, column=6, value=slab.get("thickness_mm"))
        ws.cell(row=row_idx, column=7, value=round(float(slab.get("area_m2") or 0), 3))
        ws.cell(row=row_idx, column=8, value=round(float(slab.get("concrete_m3") or 0), 3))
        ws.cell(row=row_idx, column=9, value=round(float(slab.get("shuttering_m2") or 0), 3))
        ws.cell(row=row_idx, column=10, value=slab.get("thickness_source"))
        ws.cell(row=row_idx, column=11, value=entity_conf)
        ws.cell(row=row_idx, column=12, value=stats["low_conf_entity_count"])
        ws.cell(row=row_idx, column=13, value="YES" if stats["review_required"] else "")
        ws.cell(row=row_idx, column=14, value=centroid[0] if len(centroid) > 0 else None)
        ws.cell(row=row_idx, column=15, value=centroid[1] if len(centroid) > 1 else None)
        if stats["review_required"]:
            fill = _review_fill(
                entity_conf if entity_conf is not None else float(slab.get("thickness_confidence") or 1) * 100,
                review_threshold=review_threshold,
                force_threshold=force_threshold,
            ) or FILL_REVIEW
            _highlight_row(ws, row_idx, len(headers), fill)

    total_row = len(slabs) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=round(sum(float(s.get("area_m2") or 0) for s in slabs), 3)).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=round(sum(float(s.get("concrete_m3") or 0) for s in slabs), 3)).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=round(sum(float(s.get("shuttering_m2") or 0) for s in slabs), 3)).font = Font(bold=True)
    _autosize_columns(ws)


def _write_beams_sheet(
    wb,
    beams: list[dict[str, Any]],
    *,
    review_threshold: float = DEFAULT_REVIEW_THRESHOLD,
    force_threshold: float = 60.0,
) -> None:
    if not beams:
        return
    ws = wb.create_sheet("Beams")
    headers = [
        "S.No",
        "Beam ID",
        "Component ID",
        "Layer",
        "Length (m)",
        "Width (mm)",
        "Depth (mm)",
        "Section Source",
        "Concrete (m3)",
        "Shuttering (m2)",
        "Confidence",
        "Review",
        "Centroid X (mm)",
        "Centroid Y (mm)",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_idx, beam in enumerate(beams, start=2):
        centroid = beam.get("centroid_mm") or [None, None]
        confidence = beam.get("confidence")
        review_flag = bool(beam.get("review_required")) or (
            confidence is not None and float(confidence) < review_threshold
        )
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=beam.get("beam_id"))
        ws.cell(row=row_idx, column=3, value=beam.get("component_id"))
        ws.cell(row=row_idx, column=4, value=beam.get("layer"))
        ws.cell(row=row_idx, column=5, value=beam.get("length_m"))
        ws.cell(row=row_idx, column=6, value=beam.get("width_mm"))
        ws.cell(row=row_idx, column=7, value=beam.get("depth_mm"))
        ws.cell(row=row_idx, column=8, value=beam.get("section_source"))
        ws.cell(row=row_idx, column=9, value=round(float(beam.get("concrete_m3") or 0), 3))
        ws.cell(row=row_idx, column=10, value=round(float(beam.get("shuttering_m2") or 0), 3))
        ws.cell(row=row_idx, column=11, value=confidence)
        ws.cell(row=row_idx, column=12, value="YES" if review_flag else "")
        ws.cell(row=row_idx, column=13, value=centroid[0] if len(centroid) > 0 else None)
        ws.cell(row=row_idx, column=14, value=centroid[1] if len(centroid) > 1 else None)
        conf_val = float(confidence) if confidence is not None else None
        fill = _review_fill(conf_val, review_threshold=review_threshold, force_threshold=force_threshold)
        if fill is not None:
            _highlight_row(ws, row_idx, len(headers), fill)

    total_row = len(beams) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=round(sum(float(b.get("length_m") or 0) for b in beams), 3)).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=round(sum(float(b.get("concrete_m3") or 0) for b in beams), 3)).font = Font(bold=True)
    ws.cell(row=total_row, column=10, value=round(sum(float(b.get("shuttering_m2") or 0) for b in beams), 3)).font = Font(bold=True)
    _autosize_columns(ws)


def _write_review_sheet(
    wb,
    review_data: dict[str, Any] | None,
    *,
    review_threshold: float = DEFAULT_REVIEW_THRESHOLD,
    force_threshold: float = 60.0,
) -> None:
    if not review_data:
        return
    entities = [
        e
        for e in review_data.get("entities") or []
        if e.get("classification") in ("Slab", "Beam")
    ]
    if not entities:
        return

    ws = wb.create_sheet("Review Queue")
    ws.cell(row=1, column=1, value="Estimator review — low-confidence Slab / Beam entities").font = Font(bold=True)
    ws.cell(
        row=2,
        column=1,
        value=(
            f"Highlight: amber < {review_threshold:.0f}% confidence, "
            f"orange < {force_threshold:.0f}%"
        ),
    )

    headers = [
        "S.No",
        "Entity ID",
        "Classification",
        "Confidence %",
        "Review",
        "Layer",
        "Alternatives",
        "Rule Evidence",
    ]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)

    for idx, ent in enumerate(entities, start=1):
        row_idx = header_row + idx
        confidence = ent.get("confidence")
        conf_val = float(confidence) if confidence is not None else None
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=ent.get("entity_id"))
        ws.cell(row=row_idx, column=3, value=ent.get("classification"))
        ws.cell(row=row_idx, column=4, value=confidence)
        ws.cell(
            row=row_idx,
            column=5,
            value="YES"
            if ent.get("review_required")
            or (conf_val is not None and conf_val < review_threshold)
            else "",
        )
        ws.cell(row=row_idx, column=6, value=ent.get("layer"))
        ws.cell(row=row_idx, column=7, value=", ".join(ent.get("alternatives") or []))
        ws.cell(row=row_idx, column=8, value="; ".join(ent.get("rule_evidence") or []))
        fill = _review_fill(conf_val, review_threshold=review_threshold, force_threshold=force_threshold)
        if fill is not None:
            _highlight_row(ws, row_idx, len(headers), fill)

    _autosize_columns(ws)


def _write_metrics_sheet(wb, result: dict[str, Any]) -> None:
    inf = result.get("inference_metrics") or {}
    if not inf:
        return
    ws = wb.create_sheet("Metrics")
    bold = Font(bold=True)
    row = 1
    ws.cell(row=row, column=1, value="SDIE Inference Metrics").font = bold
    row += 2

    def section(title: str) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = bold
        row += 1

    def kv(label: str, value: Any) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    section("ML split")
    kv("Split", inf.get("split", "test"))
    kv("Metric type", inf.get("metric_type"))
    row += 1

    gt = inf.get("ground_truth") or {}
    section("Accuracy evaluation")
    kv("Workbook benchmark", gt.get("workbook_benchmark"))
    kv("Train accuracy (X)", gt.get("train_accuracy_eval"))
    kv("Test accuracy (Y)", gt.get("test_accuracy_eval"))
    row += 1

    cls = inf.get("classification") or {}
    section("Classification")
    kv("Entities total", cls.get("entities_total"))
    kv("Slab+Beam entity count", cls.get("slab_beam_entity_count"))
    kv("Slab+Beam share %", cls.get("slab_beam_share_pct"))
    kv("Mean confidence (all) %", cls.get("mean_confidence_all"))
    kv("Mean confidence (Slab+Beam) %", cls.get("mean_confidence_slab_beam"))
    kv("Low confidence count", cls.get("low_confidence_count"))
    kv("Low confidence %", cls.get("low_confidence_pct"))
    kv("Review required count", cls.get("review_required_count"))
    kv("Review required %", cls.get("review_required_pct"))
    kv("Unknown count", cls.get("unknown_count"))
    kv("Unknown %", cls.get("unknown_pct"))
    kv("DeepSeek ambiguous", cls.get("ambiguous_count"))
    kv("DeepSeek updated", cls.get("deepseek_updated"))
    row += 1

    qty = inf.get("quantities") or {}
    section("Quantities")
    kv("Slab count", qty.get("slab_count"))
    kv("Slab area (m2)", qty.get("slab_area_m2"))
    kv("Slab concrete (m3)", qty.get("slab_concrete_m3"))
    kv("Beam count", qty.get("beam_count"))
    kv("Beam total length (m)", qty.get("beam_total_length_m"))
    kv("Beam concrete (m3)", qty.get("beam_concrete_m3"))
    kv("Total concrete (m3)", qty.get("total_concrete_m3"))
    kv("Total shuttering (m2)", qty.get("total_shuttering_m2"))
    row += 1

    section("Quality targets")
    targets = inf.get("quality_targets") or {}
    for key, val in targets.items():
        kv(key, val)
    kv("Overall quality pass", inf.get("quality_pass"))
    row += 1

    checks = inf.get("quality_checks") or []
    if checks:
        section("Quality checks")
        ws.cell(row=row, column=1, value="Check").font = bold
        ws.cell(row=row, column=2, value="Value").font = bold
        ws.cell(row=row, column=3, value="Target").font = bold
        ws.cell(row=row, column=4, value="Pass").font = bold
        row += 1
        for item in checks:
            ws.cell(row=row, column=1, value=item.get("check"))
            ws.cell(row=row, column=2, value=item.get("value"))
            ws.cell(row=row, column=3, value=item.get("target"))
            ws.cell(row=row, column=4, value=item.get("pass"))
            row += 1
        row += 1

    per_type = cls.get("per_type") or {}
    if per_type:
        section("Per-type confidence")
        ws.cell(row=row, column=1, value="Component Type").font = bold
        ws.cell(row=row, column=2, value="Count").font = bold
        ws.cell(row=row, column=3, value="Mean confidence %").font = bold
        ws.cell(row=row, column=4, value="Low confidence %").font = bold
        row += 1
        for ctype, stats in sorted(per_type.items()):
            if not isinstance(stats, dict):
                continue
            ws.cell(row=row, column=1, value=ctype)
            ws.cell(row=row, column=2, value=stats.get("count"))
            ws.cell(row=row, column=3, value=stats.get("mean_confidence"))
            ws.cell(row=row, column=4, value=stats.get("low_confidence_pct"))
            row += 1

    _autosize_columns(ws)


def _write_classification_sheet(wb, result: dict[str, Any]) -> None:
    notes = result.get("detection_notes") or {}
    counts = notes.get("component_type_counts") or {}
    if not counts:
        return
    ws = wb.create_sheet("Classification")
    ws.cell(row=1, column=1, value="Component Type").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Count").font = Font(bold=True)
    for row_idx, (ctype, count) in enumerate(sorted(counts.items()), start=2):
        ws.cell(row=row_idx, column=1, value=ctype)
        ws.cell(row=row_idx, column=2, value=count)
    _autosize_columns(ws)


def export_results_to_excel(
    result: dict[str, Any],
    output_path: Path,
) -> Path:
    """Write pipeline result dict to an Excel workbook."""
    _require_openpyxl()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    ctx = _load_export_context(result, output_path)
    review_threshold = ctx["review_threshold"]
    force_threshold = ctx["force_threshold"]

    wb = openpyxl.Workbook()
    _write_summary_sheet(wb, result)
    _write_metrics_sheet(wb, result)
    _write_review_sheet(
        wb,
        ctx["review_queue"],
        review_threshold=review_threshold,
        force_threshold=force_threshold,
    )
    _write_slabs_sheet(
        wb,
        result.get("slabs") or [],
        components=ctx["components"],
        review_threshold=review_threshold,
        force_threshold=force_threshold,
    )
    _write_beams_sheet(
        wb,
        result.get("beams") or [],
        review_threshold=review_threshold,
        force_threshold=force_threshold,
    )
    _write_classification_sheet(wb, result)
    wb.save(output_path)
    return output_path


def export_results_json_to_excel(json_path: Path, output_path: Path | None = None) -> Path:
    """Load a *_results.json file and export to Excel."""
    import json

    result = json.loads(json_path.read_text(encoding="utf-8"))
    out = output_path or json_path.with_name(json_path.stem.replace("_results", "") + "_quantities.xlsx")
    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")
    return export_results_to_excel(result, out)
