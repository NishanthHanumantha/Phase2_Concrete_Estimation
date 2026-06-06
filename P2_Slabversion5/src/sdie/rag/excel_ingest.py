from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore


def _num(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if str(value).startswith("#"):
            return None
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text or text.startswith("#"):
            return None
        try:
            return float(text.replace(",", ""))
        except ValueError:
            return None
    return None


def _norm(text: Any) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).strip()).lower()


def _load_workbook(path: Path):
    if openpyxl is None:
        raise RuntimeError("openpyxl is required: pip install openpyxl")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _sheet_rows(path: Path, sheet_name: str, max_row: int = 200) -> list[tuple[Any, ...]]:
    wb = _load_workbook(path)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        return [tuple(row) for row in ws.iter_rows(max_row=max_row, values_only=True)]
    finally:
        wb.close()


def _find_header_row(
    rows: list[tuple[Any, ...]],
    *,
    must_contain: tuple[str, ...],
) -> tuple[int, list[str]] | None:
    for idx, row in enumerate(rows):
        labels = [_norm(c) for c in row]
        if all(any(token in label for label in labels) for token in must_contain):
            return idx, labels
    return None


def _column_index(labels: list[str], *tokens: str) -> int | None:
    for i, label in enumerate(labels):
        if all(token in label for token in tokens):
            return i
    return None


def extract_inizio_tower_slab_floor(
    workbook: Path,
    floor_label: str,
) -> dict[str, float]:
    """Read 'Tower slab' abstract row for one floor (shuttering m2, concrete m3)."""
    rows = _sheet_rows(workbook, "Tower slab", max_row=30)
    header = _find_header_row(rows, must_contain=("description", "total"))
    if not header:
        return {}
    header_idx, labels = header
    floor_key = _norm(floor_label)
    floor_col = None
    for i, label in enumerate(labels):
        if floor_key in label or label.replace("-", "") in floor_key.replace("-", ""):
            floor_col = i
            break
    if floor_col is None:
        return {}

    out: dict[str, float] = {}
    for row in rows[header_idx + 1 : header_idx + 6]:
        desc = _norm(row[1] if len(row) > 1 else "")
        val = _num(row[floor_col] if len(row) > floor_col else None)
        if val is None:
            continue
        if "shuttering" in desc and "slab" in desc:
            out["shuttering_m2"] = val
            out["area_m2"] = val
        elif "concrete" in desc and "slab" in desc:
            out["concrete_m3"] = val
    return out


def extract_manohar_slabs(workbook: Path) -> dict[str, Any]:
    rows = _sheet_rows(workbook, "Slabs", max_row=120)
    expected: dict[str, float] = {}
    slabs: list[dict[str, Any]] = []

    for row in rows[:12]:
        cells = [c for c in row if c is not None]
        if len(cells) >= 4 and isinstance(cells[0], (int, float)):
            desc = str(cells[1]) if len(cells) > 1 else ""
            conc = _num(cells[2] if len(cells) > 2 else None)
            shut = _num(cells[3] if len(cells) > 3 else None)
            if conc and shut and "slab" in _norm(desc):
                slabs.append(
                    {
                        "slab_id": f"SUMMARY-{int(cells[0])}",
                        "name": desc,
                        "is_structural_slab": True,
                        "area_m2": shut,
                        "expected_concrete_m3": conc,
                        "expected_shuttering_m2": shut,
                        "validation_status": "from_workbook",
                    }
                )

    for row in rows:
        if len(row) < 8:
            continue
        if not isinstance(row[0], (int, float)):
            continue
        slab_id = row[1]
        if not isinstance(slab_id, str) or not re.match(r"^S\d+", slab_id.strip(), re.I):
            continue
        length = _num(row[3])
        breadth = _num(row[4])
        depth = _num(row[5])
        area = _num(row[6])
        conc = _num(row[7])
        if area is None and length and breadth:
            area = length * breadth
        if area is None:
            continue
        slabs.append(
            {
                "slab_id": slab_id.strip(),
                "name": slab_id.strip(),
                "is_structural_slab": True,
                "thickness_mm": int(depth * 1000) if depth and depth < 2 else None,
                "area_m2": area,
                "expected_concrete_m3": conc or 0.0,
                "expected_shuttering_m2": area,
                "validation_status": "from_workbook",
            }
        )

    if slabs:
        expected["area_m2"] = round(sum(s["area_m2"] for s in slabs if s["area_m2"]), 3)
        expected["concrete_m3"] = round(
            sum(s.get("expected_concrete_m3") or 0 for s in slabs), 3
        )
        expected["shuttering_m2"] = expected["area_m2"]
        expected["slab_count"] = len(slabs)

    return {"expected_total": expected, "slabs": slabs}


def extract_trust_office_slabs(workbook: Path) -> dict[str, Any]:
    rows = _sheet_rows(workbook, "Slabs", max_row=200)
    slabs: list[dict[str, Any]] = []
    expected: dict[str, float] = {}

    for row in rows:
        if len(row) < 8:
            continue
        if not isinstance(row[0], (int, float)):
            continue
        desc = row[2]
        if not isinstance(desc, str) or not desc.strip().upper().startswith("S"):
            continue
        area = _num(row[7])
        conc = _num(row[6])
        if area is None:
            continue
        slab_id = desc.strip()
        slabs.append(
            {
                "slab_id": slab_id,
                "name": slab_id,
                "is_structural_slab": True,
                "area_m2": area,
                "expected_concrete_m3": conc or 0.0,
                "expected_shuttering_m2": area,
                "validation_status": "from_workbook",
            }
        )

    for row in rows[:20]:
        row_norm = [_norm(c) for c in row]
        if "slab" in row_norm and "sqm" in row_norm:
            for val in row:
                n = _num(val)
                if n and n > 100:
                    expected.setdefault("area_m2", n)
                    expected.setdefault("shuttering_m2", n)
        if "slab" in row_norm and "cum" in row_norm:
            for val in row:
                n = _num(val)
                if n and n > 50:
                    expected.setdefault("concrete_m3", n)

    if slabs and "area_m2" not in expected:
        expected["area_m2"] = round(sum(s["area_m2"] for s in slabs), 3)
        expected["concrete_m3"] = round(
            sum(s.get("expected_concrete_m3") or 0 for s in slabs), 3
        )
        expected["shuttering_m2"] = expected["area_m2"]
    if slabs:
        expected["slab_count"] = len(slabs)

    return {"expected_total": expected, "slabs": slabs}


def build_ground_truth_from_project(
    *,
    project: dict[str, Any],
    data_source_root: Path,
    output_dir: Path,
) -> list[Path]:
    """Convert one manifest project workbook into per-drawing ground_truth JSON files."""
    workbook_rel = project.get("workbook")
    if not workbook_rel:
        return []
    workbook = data_source_root / workbook_rel
    if not workbook.is_file():
        raise FileNotFoundError(workbook)

    project_id = project["project_id"]
    written: list[Path] = []

    if project_id == "MANOHAR":
        bundle = extract_manohar_slabs(workbook)
    elif project_id == "TRUST_OFFICE":
        bundle = extract_trust_office_slabs(workbook)
    else:
        bundle = {"expected_total": {}, "slabs": []}

    for drawing in project.get("drawings", []):
        drawing_id = drawing["drawing_id"]
        dxf_rel = drawing["dxf"]
        expected = dict(bundle.get("expected_total", {}))
        slabs = list(bundle.get("slabs", []))

        if drawing.get("excel_floor") and project_id == "INIZIO":
            floor_vals = extract_inizio_tower_slab_floor(workbook, drawing["excel_floor"])
            if floor_vals:
                expected = floor_vals

        gt: dict[str, Any] = {
            "drawing_id": drawing_id,
            "project_id": project_id,
            "source_dxf": f"Data Source/{dxf_rel.replace(chr(92), '/')}",
            "estimator_workbook": f"Data Source/{workbook_rel.replace(chr(92), '/')}",
            "validation_status": "from_workbook",
            "expected_total": expected,
            "slabs": slabs if drawing.get("primary") or project_id != "INIZIO" else [],
            "regression_config": {
                "area_tolerance_pct": 5,
                "slab_count_tolerance": 10,
                "notes": "Imported from estimator Excel workbook",
            },
        }

        if drawing.get("floor_id"):
            gt["floor_id"] = drawing["floor_id"]
        if drawing.get("excel_floor"):
            gt["floor_match"] = drawing["excel_floor"]

        if not gt["expected_total"]:
            continue

        out_path = output_dir / f"{drawing_id}.json"
        out_path.write_text(json.dumps(gt, indent=2), encoding="utf-8")
        written.append(out_path)

    return written


def ingest_all_projects(
    *,
    manifest_path: Path,
    data_source_root: Path,
    output_dir: Path,
) -> dict[str, Any]:
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    output_dir.mkdir(parents=True, exist_ok=True)
    summary: dict[str, Any] = {"projects": [], "ground_truth_files": []}

    for project in manifest.get("projects", []):
        try:
            files = build_ground_truth_from_project(
                project=project,
                data_source_root=data_source_root,
                output_dir=output_dir,
            )
        except FileNotFoundError as exc:
            summary["projects"].append(
                {
                    "project_id": project.get("project_id"),
                    "status": "missing_workbook",
                    "error": str(exc),
                }
            )
            continue

        summary["projects"].append(
            {
                "project_id": project.get("project_id"),
                "status": "ok",
                "ground_truth_count": len(files),
                "workbook": project.get("workbook"),
            }
        )
        summary["ground_truth_files"].extend(str(p) for p in files)

    return summary
